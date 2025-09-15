# routes/contacts.py
from flask import Blueprint, request, jsonify
from config import contact_lists_collection
from utils.helpers import str_to_objectid
from pymongo.errors import DuplicateKeyError
import csv
import io

bp = Blueprint("contacts", __name__)

# --- Contact Lists ---

# GET all contact lists
@bp.route("/", methods=["GET"])
def get_contact_lists():
    lists = list(contact_lists_collection.find())
    result = []
    for l in lists:
        result.append({
            "id": str(l["_id"]),
            "name": l["name"],
            "contacts": l.get("contacts", [])
        })
    return jsonify(result)

# POST create new contact list
@bp.route("/", methods=["POST"])
def create_contact_list():
    data = request.json
    name = data.get("name")
    if not name:
        return jsonify({"status": "error", "message": "Missing name"}), 400

    new_list = {"name": name, "contacts": []}
    result = contact_lists_collection.insert_one(new_list)
    return jsonify({"status": "ok", "id": str(result.inserted_id)})

# DELETE a contact list
@bp.route("/<list_id>", methods=["DELETE"])
def delete_contact_list(list_id):
    list_obj_id = str_to_objectid(list_id)
    if not list_obj_id:
        return jsonify({"status": "error", "message": "Invalid ID"}), 400

    contact_lists_collection.delete_one({"_id": list_obj_id})
    return jsonify({"status": "ok"})


# --- Contacts inside a list ---

# GET contacts of a list
@bp.route("/<list_id>/contacts", methods=["GET"])
def get_contacts(list_id):
    list_obj_id = str_to_objectid(list_id)
    if not list_obj_id:
        return jsonify({"status": "error", "message": "Invalid ID"}), 400

    contact_list = contact_lists_collection.find_one({"_id": list_obj_id})
    if not contact_list:
        return jsonify({"status": "error", "message": "List not found"}), 404

    return jsonify(contact_list.get("contacts", []))

# POST add contact to a list
@bp.route("/<list_id>/contacts", methods=["POST"])
def add_contact(list_id):
    list_obj_id = str_to_objectid(list_id)
    if not list_obj_id:
        return jsonify({"status": "error", "message": "Invalid ID"}), 400

    contact = request.json
    if not contact.get("email"):
        return jsonify({"status": "error", "message": "Missing email"}), 400

    # Add contact if email doesn't exist
    contact_lists_collection.update_one(
        {"_id": list_obj_id, "contacts.email": {"$ne": contact["email"]}},
        {"$push": {"contacts": contact}}
    )
    return jsonify({"status": "ok"})

# DELETE contact from a list
@bp.route("/<list_id>/contacts/<email>", methods=["DELETE"])
def delete_contact(list_id, email):
    list_obj_id = str_to_objectid(list_id)
    if not list_obj_id:
        return jsonify({"status": "error", "message": "Invalid ID"}), 400

    contact_lists_collection.update_one(
        {"_id": list_obj_id},
        {"$pull": {"contacts": {"email": email}}}
    )
    return jsonify({"status": "ok"})



# POST import contacts from Excel to a list
# POST import contacts from Excel to a list
@bp.route("/<list_id>/contacts/import", methods=["POST"])
def import_contacts_json(list_id):
    from bson import ObjectId

    try:
        list_obj_id = str_to_objectid(list_id)
        if not list_obj_id:
            return jsonify({"status": "error", "message": "Invalid ID"}), 400

        data = request.get_json()
        if not data or "contacts" not in data:
            return jsonify({"status": "error", "message": "No contacts provided"}), 400

        contacts_to_import = data["contacts"]
        if not isinstance(contacts_to_import, list):
            return jsonify({"status": "error", "message": "Contacts should be a list"}), 400

        # Dohvati postojeće kontakte iz liste
        contact_list = contact_lists_collection.find_one({"_id": list_obj_id})
        existing_emails = set(c["email"].lower() for c in contact_list.get("contacts", []))

        # Filtriraj duplikate
        unique_contacts = []
        for c in contacts_to_import:
            email = c.get("email")
            if email and email.lower() not in existing_emails:
                unique_contacts.append({
                    "email": email.strip(),
                    "name": c.get("name", "").strip()
                })
                existing_emails.add(email.lower())

        if not unique_contacts:
            return jsonify({"status": "ok", "imported": 0, "message": "No new contacts to import"})

        # Dodaj samo unikatne kontakte
        contact_lists_collection.update_one(
            {"_id": list_obj_id},
            {"$push": {"contacts": {"$each": unique_contacts}}}
        )

        return jsonify({"status": "ok", "imported": len(unique_contacts)})

    except Exception as e:
        print("Error importing contacts JSON:", e)
        return jsonify({"status": "error", "message": "Failed to import contacts"}), 500

    from openpyxl import load_workbook
    from io import BytesIO

    try:
        list_obj_id = str_to_objectid(list_id)
        if not list_obj_id:
            return jsonify({"status": "error", "message": "Invalid ID"}), 400

        if "file" not in request.files:
            return jsonify({"status": "error", "message": "No file uploaded"}), 400

        file = request.files["file"]
        if not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
            return jsonify({"status": "error", "message": "Only Excel files are supported"}), 415

        wb = load_workbook(filename=BytesIO(file.read()))
        sheet = wb.active

        headers = [str(cell.value).strip().lower() for cell in sheet[1]]
        imported_contacts = []

        # postojeći emaili u listi
        contact_list = contact_lists_collection.find_one({"_id": list_obj_id})
        existing_emails = set(c["email"].lower() for c in contact_list.get("contacts", []))

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            email = row_dict.get("email")
            if email and email.lower() not in existing_emails:
                contact = {
                    "email": str(email).strip(),
                    "name": str(row_dict.get("name") or "").strip()
                }
                imported_contacts.append(contact)
                existing_emails.add(email.lower())  # dodaj u set da spriječiš duplikate unutar istog file-a

        if not imported_contacts:
            return jsonify({"status": "error", "message": "No new contacts to import"}), 400

        contact_lists_collection.update_one(
            {"_id": list_obj_id},
            {"$push": {"contacts": {"$each": imported_contacts}}}
        )

        return jsonify({"status": "ok", "imported": len(imported_contacts)})

    except Exception as e:
        print("Error importing Excel:", e)
        return jsonify({"status": "error", "message": "Failed to import Excel"}), 500

    from openpyxl import load_workbook
    from io import BytesIO
    from bson import ObjectId

    try:
        list_obj_id = str_to_objectid(list_id)
        if not list_obj_id:
            return jsonify({"status": "error", "message": "Invalid ID"}), 400

        if "file" not in request.files:
            return jsonify({"status": "error", "message": "No file uploaded"}), 400

        file = request.files["file"]
        if not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
            return jsonify({"status": "error", "message": "Only Excel files are supported"}), 415

        wb = load_workbook(filename=BytesIO(file.read()))
        sheet = wb.active

        headers = [str(cell.value).strip().lower() for cell in sheet[1]]
        imported_contacts = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            email = row_dict.get("email")
            if email:
                contact = {
                    "email": str(email).strip(),
                    "name": str(row_dict.get("name") or "").strip()
                }
                imported_contacts.append(contact)

        if not imported_contacts:
            return jsonify({"status": "error", "message": "No valid contacts found"}), 400

        contact_lists_collection.update_one(
            {"_id": list_obj_id},
            {"$push": {"contacts": {"$each": imported_contacts}}}
        )

        return jsonify({"status": "ok", "imported": len(imported_contacts)})

    except Exception as e:
        print("Error importing Excel:", e)
        return jsonify({"status": "error", "message": "Failed to import Excel"}), 500

    from openpyxl import load_workbook
    from io import BytesIO

    list_obj_id = str_to_objectid(list_id)
    if not list_obj_id:
        return jsonify({"status": "error", "message": "Invalid ID"}), 400

    if "file" not in request.files:
        return jsonify({"status": "error", "message": "No file uploaded"}), 400

    file = request.files["file"]

    if not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        return jsonify({"status": "error", "message": "Only Excel files are supported"}), 415

    try:
        wb = load_workbook(filename=BytesIO(file.read()))
        sheet = wb.active

        headers = [cell.value for cell in sheet[1]]  # prvi red je header
        imported_contacts = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            if row_dict.get("email"):
                contact = {
                    "email": str(row_dict.get("email")).strip(),
                    "name": str(row_dict.get("name") or "").strip()
                }
                imported_contacts.append(contact)

        if imported_contacts:
            contact_lists_collection.update_one(
                {"_id": list_obj_id},
                {"$push": {"contacts": {"$each": imported_contacts}}}
            )

        return jsonify({"status": "ok", "imported": len(imported_contacts)})

    except Exception as e:
        print("Error importing Excel:", e)
        return jsonify({"status": "error", "message": "Failed to import Excel"}), 500

